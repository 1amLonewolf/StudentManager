"""
Reports Module
Generate various reports and export data
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_student_report(student, course_fee=4000):
    """
    Generate comprehensive report for a student

    Args:
        student: Student object
        course_fee: Course fee (default 4000 KES)
    """
    report = {
        'student': {
            'name': student.name,
            'phone': student.phone,
            'email': student.email,
            'id_number': student.id_number,
            'cohort': student.cohort,
            'enrollment_date': student.enrollment_date.strftime('%d %B %Y') if student.enrollment_date else 'N/A',
            'status': student.status.capitalize()
        },
        'attendance': {
            'total_days': len(student.attendances),
            'present_days': sum(1 for a in student.attendances if a.present),
            'absent_days': sum(1 for a in student.attendances if not a.present),
            'rate': student.attendance_rate
        },
        'assessments': [],
        'performance': {
            'average_score': student.average_score,
            'modules_passed': 0,
            'modules_total': len(student.assessments),
            'grade': ''
        },
        'payments': {
            'total_paid': student.total_paid,
            'pending': 0,  # Calculated below using course_fee
            'transactions': []
        }
    }
    
    # Add assessment details
    for assessment in student.assessments:
        report['assessments'].append({
            'module': assessment.module,
            'score': assessment.score,
            'max_score': assessment.max_score,
            'percentage': assessment.percentage,
            'type': assessment.assessment_type.capitalize(),
            'date': assessment.date_taken.strftime('%d %B %Y') if assessment.date_taken else 'N/A',
            'remarks': assessment.remarks or '-'
        })
        
        if assessment.passed:
            report['performance']['modules_passed'] += 1
    
    # Determine grade
    avg = student.average_score
    if avg >= 80:
        report['performance']['grade'] = 'Distinction'
    elif avg >= 65:
        report['performance']['grade'] = 'Credit'
    elif avg >= 50:
        report['performance']['grade'] = 'Pass'
    else:
        report['performance']['grade'] = 'Fail'
    
    # Payment details
    report['payments']['pending'] = max(0, course_fee - student.total_paid)
    
    for payment in student.payments:
        report['payments']['transactions'].append({
            'amount': payment.amount,
            'date': payment.payment_date.strftime('%d %B %Y') if payment.payment_date else 'N/A',
            'method': payment.payment_method.capitalize(),
            'reference': payment.reference or '-'
        })
    
    return report


def generate_cohort_report(cohort=None):
    """
    Generate report for a specific cohort or all students
    
    Returns:
        dict with cohort statistics
    """
    from models import Student, Attendance, Assessment, Payment, db
    from sqlalchemy import func
    
    query = Student.query
    
    if cohort:
        query = query.filter_by(cohort=cohort)
    
    students = query.all()
    
    report = {
        'cohort': cohort or 'All Students',
        'total_students': len(students),
        'active': sum(1 for s in students if s.status == 'active'),
        'completed': sum(1 for s in students if s.status == 'completed'),
        'dropped': sum(1 for s in students if s.status == 'dropped'),
        'attendance_rate': 0,
        'pass_rate': 0,
        'average_score': 0,
        'revenue': 0,
        'students': []
    }
    
    if students:
        # Calculate averages
        total_attendance_rate = sum(s.attendance_rate for s in students)
        report['attendance_rate'] = round(total_attendance_rate / len(students), 1)
        
        students_with_assessments = [s for s in students if s.assessments]
        if students_with_assessments:
            passed = sum(1 for s in students_with_assessments if s.average_score >= 50)
            report['pass_rate'] = round((passed / len(students_with_assessments)) * 100, 1)
        
        total_avg_score = sum(s.average_score for s in students if s.assessments)
        students_with_scores = [s for s in students if s.assessments]
        if students_with_scores:
            report['average_score'] = round(total_avg_score / len(students_with_scores), 1)
        
        report['revenue'] = sum(s.total_paid for s in students)
        
        # Student summary list
        for student in students:
            report['students'].append({
                'id': student.id,
                'name': student.name,
                'phone': student.phone,
                'status': student.status,
                'attendance_rate': student.attendance_rate,
                'average_score': student.average_score,
                'total_paid': student.total_paid
            })
    
    return report


def export_to_excel(export_type='students'):
    """
    Export data to Excel file
    
    Args:
        export_type: 'students', 'attendance', 'assessments', 'payments'
    
    Returns:
        BytesIO object with Excel file
    """
    from models import Student, Attendance, Assessment, Payment
    
    wb = Workbook()
    ws = wb.active
    ws.title = export_type.capitalize()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    if export_type == 'students':
        headers = ['ID', 'Name', 'Phone', 'Email', 'ID Number', 'Gender', 'Cohort', 'Status', 'Enrollment Date', 'Attendance Rate', 'Avg Score', 'Total Paid']
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for student in Student.query.all():
            ws.append([
                student.id,
                student.name,
                student.phone or '',
                student.email or '',
                student.id_number or '',
                student.gender or '',
                student.cohort or '',
                student.status,
                student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else '',
                student.attendance_rate,
                student.average_score,
                student.total_paid
            ])
    
    elif export_type == 'attendance':
        headers = ['ID', 'Student', 'Date', 'Module', 'Present', 'Notes', 'Recorded At']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for record in Attendance.query.all():
            ws.append([
                record.id,
                record.student.name,
                record.date.strftime('%Y-%m-%d'),
                record.module,
                'Yes' if record.present else 'No',
                record.notes or '',
                record.recorded_at.strftime('%Y-%m-%d %H:%M') if record.recorded_at else ''
            ])
    
    elif export_type == 'assessments':
        headers = ['ID', 'Student', 'Module', 'Score', 'Max Score', 'Percentage', 'Type', 'Date', 'Passed']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for assessment in Assessment.query.all():
            ws.append([
                assessment.id,
                assessment.student.name,
                assessment.module,
                assessment.score,
                assessment.max_score,
                assessment.percentage,
                assessment.assessment_type,
                assessment.date_taken.strftime('%Y-%m-%d') if assessment.date_taken else '',
                'Yes' if assessment.passed else 'No'
            ])
    
    elif export_type == 'payments':
        headers = ['ID', 'Student', 'Amount', 'Date', 'Method', 'Reference', 'Notes', 'Recorded By']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for payment in Payment.query.all():
            ws.append([
                payment.id,
                payment.student.name,
                payment.amount,
                payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else '',
                payment.payment_method,
                payment.reference or '',
                payment.notes or '',
                payment.recorded_by or ''
            ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def generate_dashboard_data():
    """
    Generate data for dashboard charts
    
    Returns:
        dict with chart data
    """
    from models import Student, Attendance, Assessment, Payment, Course
    from sqlalchemy import func
    from datetime import timedelta
    
    # Monthly enrollment (last 6 months)
    six_months_ago = datetime.now() - timedelta(days=180)
    enrollment_data = {}
    
    for student in Student.query.filter(Student.enrollment_date >= six_months_ago).all():
        month = student.enrollment_date.strftime('%Y-%m')
        enrollment_data[month] = enrollment_data.get(month, 0) + 1
    
    # Module pass rates
    module_stats = {}
    for assessment in Assessment.query.all():
        module = assessment.module
        if module not in module_stats:
            module_stats[module] = {'passed': 0, 'total': 0}
        module_stats[module]['total'] += 1
        if assessment.passed:
            module_stats[module]['passed'] += 1
    
    module_pass_rates = {}
    for module, stats in module_stats.items():
        if stats['total'] > 0:
            module_pass_rates[module] = round((stats['passed'] / stats['total']) * 100, 1)
    
    # Revenue by month (last 6 months)
    revenue_data = {}
    for payment in Payment.query.filter(Payment.payment_date >= six_months_ago).all():
        month = payment.payment_date.strftime('%Y-%m')
        revenue_data[month] = revenue_data.get(month, 0) + payment.amount
    
    return {
        'enrollment_trend': list(enrollment_data.items()),
        'module_pass_rates': module_pass_rates,
        'revenue_trend': list(revenue_data.items())
    }
